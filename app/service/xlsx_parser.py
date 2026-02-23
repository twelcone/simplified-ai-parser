"""
XLSX to Markdown Parser

Note: The original ai-parser uses LibreOffice (soffice --headless --convert-to html)
for Excel conversion, which preserves complex formatting and shapes.

This simplified version uses openpyxl directly to create markdown tables,
which is lighter weight but may not preserve all formatting features.
For full fidelity, LibreOffice would be required.
"""

import base64
import io
import uuid
import logging
import openpyxl
from PIL import Image

logger = logging.getLogger(__name__)

SUPPORTED_IMAGE_FORMATS = ["png", "jpg", "jpeg"]


def _get_cell_value(cell) -> str:
    """
    Get cell value as string, handling None and special types.
    Matches original _trim_all_cell_whitespace behavior.
    """
    if cell.value is None:
        return ""
    value = str(cell.value)
    # Strip trailing whitespace (matching original _trim_all_cell_whitespace)
    return value.rstrip()


def _escape_markdown_table_cell(value: str) -> str:
    """Escape special characters for markdown table cells."""
    # Replace pipe characters and newlines
    value = value.replace("|", "\\|")
    value = value.replace("\n", "<br>")
    return value


def _is_supported_image_format(img_format: str) -> bool:
    """Check if image format is supported (png, jpg, jpeg only)."""
    return img_format.lower() in SUPPORTED_IMAGE_FORMATS


def _extract_images_from_sheet(ws) -> list[dict]:
    """
    Extract images from worksheet and return as base64 data URIs.
    Matches original _replace_excel_images_from_sheet behavior.
    """
    images = []

    for image in ws._images:
        try:
            img_data = image._data()

            # Get image format using PIL
            with io.BytesIO(img_data) as buffer:
                with Image.open(buffer) as pil_img:
                    img_format = pil_img.format.lower() if pil_img.format else "png"
                    width_px, height_px = pil_img.size

            # Skip unsupported formats (matching original supported_image list)
            if not _is_supported_image_format(img_format):
                logger.info(f"Skipping unsupported image format: {img_format}")
                continue

            # Convert to base64
            base64_data = base64.b64encode(img_data).decode("utf-8")
            data_uri = f"data:image/{img_format};base64,{base64_data}"

            # Get position info (matching original anchor handling)
            from_col = image.anchor._from.col
            from_row = image.anchor._from.row

            images.append({
                "data_uri": data_uri,
                "row": from_row,
                "col": from_col,
                "width": width_px,
                "height": height_px,
                "id": uuid.uuid4().hex[:8]
            })
        except Exception as e:
            logger.warning(f"Failed to extract image: {e}")
            continue

    return images


def _is_row_empty(ws, row_idx: int, min_col: int, max_col: int) -> bool:
    """
    Check if a row is entirely empty.
    Matches original _identify_rows_to_delete logic.
    """
    for col_idx in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            return False
    return True


def _is_col_empty(ws, col_idx: int, min_row: int, max_row: int) -> bool:
    """
    Check if a column is entirely empty.
    Matches original _identify_columns_to_delete logic.
    """
    for row_idx in range(min_row, max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            return False
    return True


def _get_effective_range(ws) -> tuple:
    """
    Get the effective data range, excluding empty rows/columns.
    Matches original _reduce_excel behavior.
    """
    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    # Skip leading empty rows
    while min_row <= max_row and _is_row_empty(ws, min_row, min_col, max_col):
        min_row += 1

    # Skip trailing empty rows
    while max_row >= min_row and _is_row_empty(ws, max_row, min_col, max_col):
        max_row -= 1

    # Skip leading empty columns
    while min_col <= max_col and _is_col_empty(ws, min_col, min_row, max_row):
        min_col += 1

    # Skip trailing empty columns
    while max_col >= min_col and _is_col_empty(ws, max_col, min_row, max_row):
        max_col -= 1

    return min_row, max_row, min_col, max_col


def _sheet_to_markdown_table(ws, sheet_name: str, images: list[dict]) -> str:
    """
    Convert a worksheet to markdown table format.
    Includes image references at their anchor positions.
    """
    lines = []

    # Add sheet header
    lines.append(f"## Sheet: {sheet_name}\n")

    # Get effective data range (matching original _reduce_excel behavior)
    min_row, max_row, min_col, max_col = _get_effective_range(ws)

    # Check if sheet has data
    if max_row < min_row or max_col < min_col:
        lines.append("*Empty sheet*\n")
        return "\n".join(lines)

    # Build the table
    table_rows = []
    for row_idx in range(min_row, max_row + 1):
        row_cells = []
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = _get_cell_value(cell)

            # Check if there's an image at this position
            # Note: image anchor uses 0-based indexing
            for img in images:
                if img["row"] == row_idx - 1 and img["col"] == col_idx - 1:
                    img_md = f"![image-{img['id']}]({img['data_uri']})"
                    value = f"{value} {img_md}" if value else img_md

            row_cells.append(_escape_markdown_table_cell(value))
        table_rows.append(row_cells)

    if not table_rows:
        lines.append("*Empty sheet*\n")
        return "\n".join(lines)

    # Create markdown table
    # Header row (first data row becomes header)
    header = table_rows[0]
    lines.append("| " + " | ".join(header) + " |")

    # Separator row
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")

    # Data rows
    for row in table_rows[1:]:
        # Ensure row has same number of columns as header
        while len(row) < len(header):
            row.append("")
        lines.append("| " + " | ".join(row[:len(header)]) + " |")

    lines.append("")  # Empty line after table
    return "\n".join(lines)


def parse_xlsx_to_markdown(file_path: str) -> str:
    """
    Convert an Excel file (xlsx/xlsm) to Markdown format.

    Note: This simplified version uses openpyxl to directly read cells
    and create markdown tables. The original ai-parser uses LibreOffice
    for HTML conversion which preserves more complex formatting.

    Processing steps (matching original where applicable):
    1. Load workbook with data_only=True (get computed values)
    2. For each sheet:
       - Extract images with base64 encoding
       - Skip empty rows/columns (like original _reduce_excel)
       - Convert to markdown table format

    Args:
        file_path: Path to the Excel file

    Returns:
        Markdown content as string with tables for each sheet
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    markdown_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract images (matching original _replace_excel_images_from_sheet)
        images = _extract_images_from_sheet(ws)

        # Convert sheet to markdown table
        sheet_md = _sheet_to_markdown_table(ws, sheet_name, images)
        markdown_parts.append(sheet_md)

    wb.close()

    return "\n".join(markdown_parts)
